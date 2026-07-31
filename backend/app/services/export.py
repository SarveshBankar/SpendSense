import csv
import io
from datetime import datetime

from fastapi import HTTPException, status
from sqlalchemy.orm import Session

from app.models.user import User
from app.repositories.transaction import TransactionRepository
from app.services.analytics import AnalyticsService


class ExportService:
    def __init__(self, db: Session):
        self.db = db
        self.repo = TransactionRepository()
        self.analytics = AnalyticsService(db)

    def export_csv(self, current_user: User, month: int | None = None, year: int | None = None) -> str:
        txs = self._filtered(current_user, month, year)
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["Date", "Description", "Amount", "Type", "Category", "Merchant", "Payment Mode", "Balance"])
        for t in txs:
            writer.writerow([t.date, t.description, t.amount, t.transaction_type,
                             t.category or "", t.merchant or "", t.payment_mode or "", t.balance or ""])
        return output.getvalue()

    def export_excel(self, current_user: User, month: int | None = None, year: int | None = None) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = Workbook()
        # Sheet 1: Transactions
        ws1 = wb.active
        ws1.title = "Transactions"
        headers = ["Date", "Description", "Amount", "Type", "Category", "Merchant", "Payment Mode", "Balance"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        for col, h in enumerate(headers, 1):
            c = ws1.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center")

        txs = self._filtered(current_user, month, year)
        for row_idx, t in enumerate(txs, 2):
            ws1.cell(row=row_idx, column=1, value=t.date)
            ws1.cell(row=row_idx, column=2, value=t.description)
            ws1.cell(row=row_idx, column=3, value=t.amount)
            ws1.cell(row=row_idx, column=4, value=t.transaction_type)
            ws1.cell(row=row_idx, column=5, value=t.category or "")
            ws1.cell(row=row_idx, column=6, value=t.merchant or "")
            ws1.cell(row=row_idx, column=7, value=t.payment_mode or "")
            ws1.cell(row=row_idx, column=8, value=t.balance or "")

        for col in range(1, 9):
            ws1.column_dimensions[chr(64 + col)].width = 22

        # Sheet 2: Report Summary
        ws2 = wb.create_sheet("Summary")
        income = [t for t in txs if t.transaction_type == "credit"]
        expense = [t for t in txs if t.transaction_type == "debit"]
        total_income = sum(t.amount for t in income)
        total_expense = sum(t.amount for t in expense)
        net = total_income - total_expense
        ws2.cell(row=1, column=1, value="Metric").font = header_font
        ws2.cell(row=1, column=2, value="Value").font = header_font
        ws2.cell(row=1, column=1).fill = header_fill
        ws2.cell(row=1, column=2).fill = header_fill
        metrics = [
            ("Total Transactions", len(txs)),
            ("Total Income", round(total_income, 2)),
            ("Total Expense", round(total_expense, 2)),
            ("Net Savings", round(net, 2)),
            ("Savings Rate %", round((net / total_income * 100), 1) if total_income > 0 else 0),
            ("Generated", datetime.utcnow().isoformat()),
        ]
        for i, (k, v) in enumerate(metrics, 2):
            ws2.cell(row=i, column=1, value=k)
            ws2.cell(row=i, column=2, value=v)
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 30

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    def export_pdf(self, current_user: User, month: int | None = None, year: int | None = None) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, mm
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                rightMargin=20*mm, leftMargin=20*mm,
                                topMargin=20*mm, bottomMargin=20*mm)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=22, spaceAfter=6)
        heading = ParagraphStyle("Heading2", parent=styles["Heading2"], fontSize=14, spaceBefore=12, spaceAfter=6)
        normal = styles["Normal"]
        small = ParagraphStyle("Small", parent=normal, fontSize=9, spaceAfter=4)

        elements = []

        # Title
        elements.append(Paragraph(f"SpendSense Financial Report", title_style))
        elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", small))
        elements.append(Paragraph(f"User: {current_user.full_name} ({current_user.email})", small))
        elements.append(Spacer(1, 0.2 * inch))

        txs = self._filtered(current_user, month, year)
        income = [t for t in txs if t.transaction_type == "credit"]
        expense = [t for t in txs if t.transaction_type == "debit"]
        total_income = sum(t.amount for t in income)
        total_expense = sum(t.amount for t in expense)
        net = total_income - total_expense
        sr = round((net / total_income * 100), 1) if total_income > 0 else 0
        cat_pct = round((len([t for t in txs if t.category]) / len(txs) * 100), 0) if txs else 0

        # Summary table
        elements.append(Paragraph("Financial Summary", heading))
        summary_data = [
            ["Metric", "Value"],
            ["Period", f"{f'{year}' if year else 'All time'} {f'- Month {month}' if month else ''}"],
            ["Total Transactions", str(len(txs))],
            ["Total Income", f"₹{total_income:,.2f}"],
            ["Total Expense", f"₹{total_expense:,.2f}"],
            ["Net Savings", f"₹{net:,.2f}"],
            ["Savings Rate", f"{sr}%"],
            ["Categorized", f"{cat_pct:.0f}%"],
        ]
        t = Table(summary_data, colWidths=[150, 200])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#10B981")),
            ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
            ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 0.2 * inch))

        # Category breakdown
        cat_spend: dict[str, float] = {}
        for t in expense:
            cat_spend[t.category or "Others"] = cat_spend.get(t.category or "Others", 0) + t.amount
        cat_sorted = sorted(cat_spend.items(), key=lambda x: -x[1])
        if cat_sorted:
            elements.append(Paragraph("Category Breakdown", heading))
            cat_data = [["Category", "Amount"]]
            for cat, amt in cat_sorted:
                cat_data.append([cat, f"₹{amt:,.2f}"])
            t2 = Table(cat_data, colWidths=[200, 150])
            t2.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#6366F1")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
                ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            elements.append(t2)
            elements.append(Spacer(1, 0.2 * inch))

        # Health score & insights
        stats = {"savings_rate": sr, "expense_ratio": round((total_expense / total_income * 100), 1) if total_income > 0 else 0,
                 "categorized_pct": cat_pct, "total_transactions": len(txs),
                 "total_income": total_income, "total_expense": total_expense}
        score = self.analytics._compute_health_score(stats)
        elements.append(Paragraph(f"Financial Health Score: {score['score']}/100", heading))
        elements.append(Spacer(1, 0.1 * inch))

        # Recommendations
        recs = self.analytics._generate_recommendations(stats, [])
        if recs:
            elements.append(Paragraph("Recommendations", heading))
            for r in recs:
                elements.append(Paragraph(f"- {r['description']}", small))

        doc.build(elements)
        buf.seek(0)
        return buf.getvalue()

    def _filtered(self, current_user: User, month: int | None = None, year: int | None = None):
        txs = self.repo.list_by_user(self.db, current_user.id)
        if month is not None:
            txs = [t for t in txs if int(t.date[5:7]) == month]
        if year is not None:
            txs = [t for t in txs if int(t.date[:4]) == year]
        return txs
